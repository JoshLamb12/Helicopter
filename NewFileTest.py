from multiprocessing.sharedctypes import Value
import openpyxl as xl
import csv
from tkinter import filedialog
import glob
  

#////////////////////////////////FILE EXPLORER/////////////////////////////////////////////////
# filename = filedialog.askopenfilename(initialdir = "/",title = "Select a File",filetypes = (("Excel Files","*.xlsx*"),("all files","*.*")))
# print(filename)

x=65

path = "E:/Coding/Excel"
allfilenames = glob.glob(path + "/*.xlsx") #Determines the Folder our program will read excel documents
numFiles=len(allfilenames) #Determines the number of Files located in the designated folder

print(chr(65)+str(1))

#//////////////////////////////WORKBOOK LOAD/////////////////////////////////////////////////
names={}  #Creates empty dictionary
workbooks = {}
sheets={}




wb_full = xl.Workbook() #Creates Compilation List
sheet_full = wb_full.active
print(numFiles)

for i in range(0,numFiles):
    
    names["filename{0}".format(i)] = allfilenames[i] #Assigns Dictionary names to dictionary values for each file locacted on Path
    workbooks["wb_{0}".format(i)] = xl.load_workbook(names["filename{0}".format(i)])
    sheets["sheet_{0}".format(i)] = workbooks["wb_{0}".format(i)].active
    row = 255

    for z in range(1,row):
        f = sheets["sheet_{0}".format(i)]
        sheet_full[chr(x)+str(z)].value = f["EK"+str(z)].value
    print (x)
    print(chr(x))
    x = x+1

print(names)

# sheet_0 = sheets["sheet_0"]
# sheet_1 = sheets["sheet_1"]
# sheet_2 = sheets["sheet_2"]
# sheet_3 = sheets["sheet_3"]




# wb_0 = xl.load_workbook(names["filename0"]) #Opens First excel file on Path
# sheet_0 = wb_0.active
# wb_1 = xl.load_workbook(names["filename1"]) #Opens Second excel file on Path
# sheet_1 = wb_1.active
# wb_2 = xl.load_workbook(names["filename2"]) #Opens Third excel file on Path
# sheet_2 = wb_2.active
# wb_3 = xl.load_workbook(names["filename3"]) #Opens Fourth excel file on Path
# sheet_3 = wb_3.active

# row = max(sheet_0.max_row,sheet_1.max_row,sheet_2.max_row,sheet_3.max_row) #Determines Maximum rows between all sheets

# for r in range (1,row):
#     Visibility = sheet_1.cell(row = r, column= 141)
#     c = sheet_full.cell(row, column = 1)
#     c.value = sheet_1.cell(row, column = 141)
#     print (c.value)
#     #print(Visibility)
#     #print(Visibility.value)

# for i in range(1,row):   #Assigns columns from data sheet to columns of master sheet
#     sheet_full['A'+str(i)].value = sheet_0['A'+str(i)].value
#     sheet_full['A'+str(i)].value = sheet_1['A'+str(i)].value
#     sheet_full['C'+str(i)].value = sheet_2['A'+str(i)].value
#     sheet_full['D'+str(i)].value = sheet_3['A'+str(i)].value


for i in range(0,numFiles):
    sheet_full.insert_cols(1+(2*i))

#//////////////////////////////Name Creation//////////////////////////////////////////////////
# cleaned_string_index = filename1.find(".")
# cleaned_string = filename1[0:cleaned_string_index+15]

#/////////////////////////////WORKBOOK SAVE///////////////////////////////////////////////////
wb_full.save( "E:\\Coding\\TIMEONLY.xlsx")
      
      
      
                                                                                                  
